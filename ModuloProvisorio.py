import openpyxl
import json
from re import sub

def calculo_porcentaje(total, porciento):
    """Calcula el PORCIENTO x (x%) de un numero TOTAL."""
    return round(total*porciento/100, 2)


class ModuloProvisorio:
    """Solo es obligatorio la planilla de Provisorios, mientras la planilla de Caducos es opcional"""
    def __init__(self, rutaProvisorio=0, rutaCaduco=0):
        with open('configuracion.json', 'r') as file:
            self.config = json.load(file)
        self.premio, self.recaudacion, self.juego = self.__getRecPrem(rutaProvisorio)
        self.caduco = self.__getCaduco(rutaCaduco)
        self.cargo_automatizacion = 0
        self.comision = 0
        self.retencionIB = 0

    def __str__(self):
        cadena = u"Juego: " +str(self.juego)
        cadena +="\nRecaudación: $" +str(self.recaudacion)
        cadena +="\nPremio: $" +str(self.premio)
        cadena +="\nCargo Automatizacion: $" +str(self.cargo_automatizacion)
        cadena +="\nComision: $" +str(self.comision)
        cadena +="\nRetencion Ingresos Brutos: $" +str(self.retencionIB)
        cadena +="\nCaducos: $" +str(self.caduco)
        return cadena

    def calculo_info(self):
        """Cacula los valores de las cuentas segun los porcentajas dados en el archivo configuracion.json"""
        self.cargo_automatizacion = calculo_porcentaje(self.recaudacion, self.config['TOMBOLA']['CARGO_AUTOMATIZACION'])
        self.comision = calculo_porcentaje(self.recaudacion, self.config['TOMBOLA']['COMISION'])
        self.retencionIB = calculo_porcentaje((self.comision-self.cargo_automatizacion), self.config['TOMBOLA']['RETENCION_INGRESO_BRUTO'])

    def __getCaduco(self,rutaCaduco):
        """Si el archivo de Caducos no fue pasado al instanciar la clase, se toma por defaul 0, o sea no es oblicagorio subir el archivo"""
        if rutaCaduco == 0: return 0
        else:
            excel_document2 = openpyxl.load_workbook(rutaCaduco)
            hoja2 = excel_document2.get_sheet_by_name('Control de Caducos de Agentes -')
            return hoja2.cell(row=hoja2.max_row,column=9).value

    def __getRecPrem(self, rutaProvisorio):
        """Extrae los valores de Recaudacion y Premio, segun la planilla que se paso al instanciar la clase"""
        if rutaProvisorio == 0: return 0, 0
        excel_document = openpyxl.load_workbook(rutaProvisorio)
        hoja = excel_document.get_sheet_by_name('Resultado Provisorio del Sorteo')
        recaudacion_aux = sub(r'[^\d,]', '', hoja['C5'].value)
        premio_aux = sub(r'[^\d,]', '', hoja['C8'].value)
        juego_aux = hoja['C3'].value
        return float(premio_aux.replace(',','.')), float(recaudacion_aux.replace(',','.')), juego_aux
